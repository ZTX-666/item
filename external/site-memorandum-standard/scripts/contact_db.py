# -*- coding: utf-8 -*-
"""Shared contact registry for site-memorandum-standard (Excel .xlsx or legacy JSON)."""
from __future__ import annotations

import json
import re
import shutil
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable


def skill_root() -> Path:
    return Path(__file__).resolve().parents[1]


def default_excel_path() -> Path:
    return skill_root() / "data" / "contacts.xlsx"


def default_json_legacy_path() -> Path:
    return skill_root() / "data" / "contacts.json"


def default_db_path() -> Path:
    """Primary contacts store used by CLI and Word generator."""
    return default_excel_path()


EXCEL_COLUMNS = [
    "abbrev",
    "name",
    "name_zh",
    "title",
    "company",
    "aliases",
    "signature_image",
]


def parse_aliases(s: str | None) -> list[str]:
    if not s:
        return []
    return [x.strip() for x in re.split(r"[,，;；、\s]+", s) if x.strip()]


@dataclass
class Contact:
    abbrev: str
    #: English (roman) name for memos and matching
    name: str
    #: 中文姓名：仅用于对话与 `find()` 检索，**不出现在 Site Memo 正文**（函件姓名一律用 `name` 英文）
    name_zh: str = ""
    title: str = ""
    company: str = ""
    aliases: list[str] = field(default_factory=list)
    #: E-signature image (PNG/JPG) for memos; absolute path or path relative to skill root, e.g. "assets/signatures/xsy.png"
    signature_image: str = ""

    def memo_english_name(self) -> str:
        """Name as it must appear on Site Memo (English `name` only)."""
        en = (self.name or "").strip()
        return en or self.abbrev

    def to_memo_line(self, *, style: str = "name_title") -> str:
        """Single line for To / Copy To / Attn. — English name only (never `name_zh`)."""
        person = self.memo_english_name()
        tit = (self.title or "").strip()
        comp = (self.company or "").strip()
        if style == "name_only":
            return person or self.abbrev
        if comp:
            if tit:
                return f"{comp} — {person} ({tit})"
            return f"{comp} — {person}"
        if tit:
            return f"{person} ({tit})"
        return person or self.abbrev

    @staticmethod
    def from_dict(d: dict[str, Any]) -> Contact:
        return Contact(
            abbrev=str(d.get("abbrev", "")).strip(),
            name=str(d.get("name", "")).strip(),
            name_zh=str(d.get("name_zh", "") or "").strip(),
            title=str(d.get("title", "") or "").strip(),
            company=str(d.get("company", "") or "").strip(),
            aliases=[str(x).strip() for x in (d.get("aliases") or []) if str(x).strip()],
            signature_image=str(d.get("signature_image", "") or "").strip(),
        )


def resolve_media_path(skill_root: Path, stored: str | None) -> Path | None:
    """
    Turn a stored path string into an absolute Path if the file exists.
    Relative paths are resolved against the skill root (parent of /scripts).
    """
    s = (stored or "").strip()
    if not s:
        return None
    p = Path(s)
    if not p.is_absolute():
        p = skill_root / p
    try:
        p = p.resolve()
    except OSError:
        return None
    return p if p.is_file() else None


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _load_json(path: Path) -> tuple[int, list[Contact]]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    version = int(raw.get("version", 1))
    contacts = [Contact.from_dict(x) for x in raw.get("contacts", [])]
    return version, contacts


def _save_json(path: Path, version: int, contacts: list[Contact]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"version": version, "contacts": [asdict(c) for c in contacts]}
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _load_excel(path: Path) -> tuple[int, list[Contact]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return 1, []
    header = [_cell_str(c).lower().replace(" ", "_") for c in rows[0]]
    col_index: dict[str, int] = {}
    for i, h in enumerate(header):
        if h in EXCEL_COLUMNS:
            col_index[h] = i

    contacts: list[Contact] = []
    for row in rows[1:]:
        if not row:
            continue
        row = list(row) + [None] * len(EXCEL_COLUMNS)

        def get(col: str) -> str:
            if col not in col_index:
                return ""
            idx = col_index[col]
            if idx >= len(row):
                return ""
            return _cell_str(row[idx])

        abbrev = get("abbrev")
        if not abbrev:
            continue
        d = {
            "abbrev": abbrev,
            "name": get("name"),
            "name_zh": get("name_zh"),
            "title": get("title"),
            "company": get("company"),
            "aliases": parse_aliases(get("aliases")),
            "signature_image": get("signature_image"),
        }
        contacts.append(Contact.from_dict(d))
    return 1, contacts


def _save_excel(path: Path, version: int, contacts: list[Contact]) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "contacts"
    ws.append(EXCEL_COLUMNS)
    for c in contacts:
        ws.append(
            [
                c.abbrev,
                c.name,
                c.name_zh,
                c.title,
                c.company,
                ",".join(c.aliases) if c.aliases else "",
                c.signature_image,
            ]
        )
    widths = (12, 22, 14, 28, 32, 24, 56)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(path)


def _norm_key(s: str) -> str:
    return re.sub(r"\s+", "", s).casefold()


class ContactDB:
    def __init__(self, path: Path | None = None) -> None:
        self.path = path if path is not None else default_excel_path()
        self.version = 1
        self.contacts: list[Contact] = []

    def load(self) -> ContactDB:
        excel = self.path if self.path.suffix.lower() == ".xlsx" else None
        json_file = self.path if self.path.suffix.lower() == ".json" else None

        if excel and self.path.is_file():
            self.version, self.contacts = _load_excel(self.path)
            return self

        if json_file and self.path.is_file():
            self.version, self.contacts = _load_json(self.path)
            return self

        # Default .xlsx path but file missing: migrate legacy JSON once
        if excel and self.path == default_excel_path():
            legacy = default_json_legacy_path()
            if legacy.is_file():
                self.version, self.contacts = _load_json(legacy)
                _save_excel(self.path, self.version, self.contacts)
                bak = legacy.with_suffix(".json.bak")
                try:
                    shutil.copy2(legacy, bak)
                    legacy.unlink()
                except OSError:
                    pass
                return self

        self.contacts = []
        self.version = 1
        return self

    def save(self) -> None:
        if self.path.suffix.lower() == ".xlsx":
            _save_excel(self.path, self.version, self.contacts)
        elif self.path.suffix.lower() == ".json":
            _save_json(self.path, self.version, self.contacts)
        else:
            raise ValueError(f"Unsupported contacts file type: {self.path}")

    def _index_by_abbrev(self) -> dict[str, Contact]:
        out: dict[str, Contact] = {}
        for c in self.contacts:
            k = _norm_key(c.abbrev)
            if k:
                out[k] = c
        return out

    def find(self, token: str) -> Contact | None:
        """Match abbrev (primary), exact name, or alias. Token may include spaces."""
        t = token.strip()
        if not t:
            return None
        by_abbrev = self._index_by_abbrev()
        hit = by_abbrev.get(_norm_key(t))
        if hit:
            return hit
        nt = _norm_key(t)
        for c in self.contacts:
            if _norm_key(c.name) == nt:
                return c
            if c.name_zh and _norm_key(c.name_zh) == nt:
                return c
            for a in c.aliases:
                if _norm_key(a) == nt:
                    return c
        return None

    def upsert(self, contact: Contact) -> tuple[Contact, str]:
        """Insert or replace by abbrev (case-insensitive). Returns (contact, action)."""
        if not contact.abbrev:
            raise ValueError("abbrev is required")
        key = _norm_key(contact.abbrev)
        for i, c in enumerate(self.contacts):
            if _norm_key(c.abbrev) == key:
                self.contacts[i] = contact
                return contact, "updated"
        self.contacts.append(contact)
        return contact, "added"

    def remove(self, abbrev: str) -> bool:
        key = _norm_key(abbrev)
        before = len(self.contacts)
        self.contacts = [c for c in self.contacts if _norm_key(c.abbrev) != key]
        return len(self.contacts) < before

    def resolve_tokens(self, tokens: Iterable[str], *, style: str = "name_title", sep: str = "\n") -> str:
        """Resolve comma/semicolon/Chinese顿号-separated list of abbrevs/names."""
        raw = " ".join(tokens).strip()
        if not raw:
            return ""
        parts = re.split(r"[,，;；、\s]+", raw)
        lines: list[str] = []
        for p in parts:
            p = p.strip()
            if not p:
                continue
            c = self.find(p)
            if not c:
                raise LookupError(f"Unknown contact token: {p!r}")
            lines.append(c.to_memo_line(style=style))
        return sep.join(lines)
